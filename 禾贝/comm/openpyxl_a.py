import openpyxl
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 背景色设置函数（注意大小写）
from xird_a import xlrd_rb

class open_wb():
    '''
    测试报告写入exlex
    '''
    def __init__(self,path):
        self.path = path



    def write(self,row,value,value1,sheet='Sheet1'):
        table = openpyxl.load_workbook(self.path)
        sheet = table[sheet]
        font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color='B50007')
        sheet.cell(row=row+1, column=6, value=value).font = font
        sheet.cell(row=row+1, column=7, value=value1).font = font
        table.save(self.path)

